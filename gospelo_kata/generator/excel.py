"""Excel document generator using openpyxl.

Generates structured Excel workbooks from data.
Requires openpyxl: pip install openpyxl

Falls back gracefully if openpyxl is not installed.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _require_openpyxl() -> None:
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel generation. "
            "Install it with: pip install openpyxl"
        )


# -- Style constants --

_HEADER_FONT = None
_HEADER_FILL = None
_HEADER_ALIGNMENT = None
_THIN_BORDER = None


def _init_styles() -> None:
    global _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGNMENT, _THIN_BORDER
    if _HEADER_FONT is not None:
        return
    _HEADER_FONT = Font(bold=True, size=11)
    _HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style="thin")
    _THIN_BORDER = Border(left=side, right=side, top=side, bottom=side)


def _style_header_row(ws: Any, row: int, col_count: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _style_data_cell(ws: Any, row: int, col: int) -> None:
    """Apply data cell styling."""
    cell = ws.cell(row=row, column=col)
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def generate_checklist_excel(data: dict[str, Any], output_path: str | Path) -> Path:
    """Generate a checklist Excel workbook.

    Args:
        data: Parsed checklist data.
        output_path: Output .xlsx path.

    Returns:
        Path to the generated file.
    """
    _require_openpyxl()
    _init_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"

    # Header
    headers = ["ID", "Suite", "Item", "Item (JA)", "Target", "Requirements", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    # Data
    row = 2
    for cat in data.get("categories", []):
        for item in cat.get("items", []):
            item_id = item["id"]
            name_slug = item["name"].lower().replace(" ", "_")
            id_parts = item_id.split("-")
            prefix = f"{int(id_parts[0]):02d}" if len(id_parts) == 1 else f"{int(id_parts[0]):02d}-{id_parts[1]}"
            suite_name = f"{prefix}_{name_slug}"

            values = [
                item_id,
                suite_name,
                item["name"],
                item.get("name_ja", ""),
                item.get("target", ""),
                item.get("requirements", ""),
                item.get("status", "draft"),
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row, column=col, value=val)
                _style_data_cell(ws, row, col)
            row += 1

    # Column widths
    widths = [6, 30, 25, 25, 25, 50, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    out = Path(output_path)
    wb.save(out)
    return out


def generate_test_spec_excel(
    spec_data: dict[str, Any],
    prereq_data: dict[str, Any] | None = None,
    output_path: str | Path = "test_spec.xlsx",
    evidence_rounds: list[dict[str, Any]] | None = None,
) -> Path:
    """Generate a test specification Excel workbook.

    Creates up to 3 sheets:
    - Summary: test metadata and category summary
    - Prerequisites: test prerequisites (if prereq_data provided)
    - TestCases: test case list with result columns

    Args:
        spec_data: Parsed test spec data.
        prereq_data: Optional parsed test prereq data.
        output_path: Output .xlsx path.
        evidence_rounds: Optional list of evidence JSON dicts for result columns.

    Returns:
        Path to the generated file.
    """
    _require_openpyxl()
    _init_styles()

    wb = openpyxl.Workbook()

    # -- Summary sheet --
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, spec_data, prereq_data, evidence_rounds)

    # -- Prerequisites sheet --
    if prereq_data:
        ws_prereq = wb.create_sheet("Prerequisites")
        _build_prereq_sheet(ws_prereq, prereq_data)

    # -- TestCases sheet --
    ws_cases = wb.create_sheet("TestCases")
    _build_testcases_sheet(ws_cases, spec_data, evidence_rounds)

    out = Path(output_path)
    wb.save(out)
    return out


def _build_summary_sheet(
    ws: Any,
    spec_data: dict[str, Any],
    prereq_data: dict[str, Any] | None,
    evidence_rounds: list[dict[str, Any]] | None,
) -> None:
    """Build the Summary sheet."""
    test_name = spec_data.get("test_name", prereq_data.get("test_name", "Test") if prereq_data else "Test")
    ws.cell(row=1, column=1, value=test_name).font = Font(bold=True, size=14)

    row = 3
    if prereq_data:
        summary = prereq_data.get("summary", {})
        meta = [
            ("Project", summary.get("project", "")),
            ("Target", summary.get("target", "")),
            ("DB", summary.get("db", "")),
            ("Client", summary.get("client", "")),
            ("Environment", summary.get("test_environment", "")),
        ]
        for key, val in meta:
            if val:
                ws.cell(row=row, column=1, value=key).font = Font(bold=True)
                ws.cell(row=row, column=2, value=val)
                row += 1
        row += 1

    # Category summary
    test_cases = spec_data.get("test_cases", [])
    categories: dict[str, int] = {}
    for case in test_cases:
        cat = case.get("category", "Uncategorized")
        categories[cat] = categories.get(cat, 0) + 1

    ws.cell(row=row, column=1, value="Category").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Cases").font = Font(bold=True)
    if prereq_data:
        ws.cell(row=row, column=3, value="Risk").font = Font(bold=True)
    _style_header_row(ws, row, 3 if prereq_data else 2)
    row += 1

    risk_map = prereq_data.get("risk_map", {}) if prereq_data else {}
    for cat, count in categories.items():
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=count)
        if risk_map:
            ws.cell(row=row, column=3, value=risk_map.get(cat, ""))
        row += 1

    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(test_cases)).font = Font(bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50


def _build_prereq_sheet(ws: Any, prereq_data: dict[str, Any]) -> None:
    """Build the Prerequisites sheet."""
    row = 1

    sections = [
        ("Background", "background", "key_value"),
        ("Test Perspectives", "perspectives", "name_desc"),
        ("Constraints", "constraints", "list"),
        ("Environment", "environment", "key_value"),
        ("Test Data", "test_data", "key_value"),
        ("References", "references", "name_path"),
    ]

    for title, key, fmt in sections:
        items = prereq_data.get(key, [])
        if not items:
            continue

        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1

        if fmt == "key_value":
            for item in items:
                ws.cell(row=row, column=1, value=item["key"])
                ws.cell(row=row, column=2, value=item["value"])
                row += 1
        elif fmt == "name_desc":
            for item in items:
                ws.cell(row=row, column=1, value=item["name"])
                ws.cell(row=row, column=2, value=item["description"])
                row += 1
        elif fmt == "name_path":
            for item in items:
                ws.cell(row=row, column=1, value=item["name"])
                ws.cell(row=row, column=2, value=item["path"])
                row += 1
        elif fmt == "list":
            for item in items:
                ws.cell(row=row, column=1, value=item)
                row += 1

        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def _build_testcases_sheet(
    ws: Any,
    spec_data: dict[str, Any],
    evidence_rounds: list[dict[str, Any]] | None,
) -> None:
    """Build the TestCases sheet."""
    headers = ["ID", "Category", "Description", "Expected Result", "Priority"]

    num_rounds = len(evidence_rounds) if evidence_rounds else 0
    for i in range(1, num_rounds + 1):
        headers.extend([f"R{i} Status", f"R{i} Detail"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    test_cases = spec_data.get("test_cases", [])
    for row_idx, case in enumerate(test_cases, 2):
        values = [
            case["test_id"],
            case.get("category", ""),
            case["description"],
            case["expected_result"],
            case.get("priority", ""),
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)
            _style_data_cell(ws, row_idx, col)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10


def generate(
    data: dict[str, Any],
    output_path: str | Path,
    doc_type: str | None = None,
    **kwargs: Any,
) -> Path:
    """Auto-detect document type and generate Excel.

    Args:
        data: Parsed JSON data.
        output_path: Output .xlsx path.
        doc_type: Explicit document type. Auto-detected if None.
        **kwargs: Additional arguments (prereq_data, evidence_rounds).

    Returns:
        Path to the generated file.
    """
    if doc_type is None:
        from ..validator import detect_schema
        doc_type = detect_schema(data)

    if doc_type == "checklist":
        return generate_checklist_excel(data, output_path)
    elif doc_type in ("test_spec", "test_prereq"):
        prereq_data = kwargs.get("prereq_data")
        evidence_rounds = kwargs.get("evidence_rounds")
        return generate_test_spec_excel(
            data, prereq_data=prereq_data,
            output_path=output_path, evidence_rounds=evidence_rounds,
        )
    else:
        # Fallback: treat as checklist-like
        return generate_checklist_excel(data, output_path)
